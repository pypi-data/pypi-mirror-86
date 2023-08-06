#!/usr/bin/env python
# -*- coding: utf-8 -*-

#    This file is part of delimited2fixedwidth and is MIT-licensed.

import argparse
import codecs
import csv
import datetime
import logging
import os
import pathlib
import re
import shutil
import sys
from locale import LC_NUMERIC, atof, setlocale

from openpyxl import load_workbook

SUPPORTED_OUTPUT_FORMATS = None


def define_supported_output_formats():
    global SUPPORTED_OUTPUT_FORMATS
    SUPPORTED_OUTPUT_FORMATS = [
        "Integer",
        "Decimal",
        "Keep numeric",
        "Time",
        "Text",
    ]
    delimiters = ["/", "-", ".", ""]
    date_patterns = ["YYYYMMDD"]
    for d in delimiters:
        date_patterns.append("MM{0}DD{0}YYYY".format(d))
        date_patterns.append("DD{0}MM{0}YYYY".format(d))
    for p1 in date_patterns:
        for p2 in date_patterns:
            SUPPORTED_OUTPUT_FORMATS.append("Date ({0} to {1})".format(p2, p1))


def write_output_file(output_content, output_file):
    with open(output_file, "w") as ofile:
        ofile.write("\n".join(output_content))


def pad_output_value(val, output_format, length):
    if (
        output_format
        in (
            "Integer",
            "Decimal",
            "Keep numeric",
            "Time",
        )
        or output_format.startswith("Date (")
    ):
        # Numbers get padded with 0's added in front (to the left)
        val = str(val).zfill(length)
    else:
        # Strings get padded with spaces added to the right
        format_template = "{:<%d}" % length
        val = format_template.format(val)
    return val


def determine_date_delimiters(output_format):
    # Identify the delimiter in the input format
    supported_delimiters = ("/", "-", ".")
    in_delim = output_format[8]
    if in_delim not in supported_delimiters:
        in_delim = ""

    # Identify the delimiter in the output format
    out_delim = output_format[-4]
    if out_delim == "Y":
        out_delim = output_format[-6]
    if out_delim not in supported_delimiters:
        out_delim = ""
    return (in_delim, out_delim)


def parse_input_date(value, output_format, in_delim):
    m = None
    if output_format.startswith("Date (DD"):
        m = re.match(
            r"([0123]?\d){in_delim}([01]?\d){in_delim}(\d{{4}})".format(
                in_delim=in_delim
            ),
            value,
        )
    elif output_format.startswith("Date (MM"):
        m = re.match(
            r"([01]?\d){in_delim}([0123]?\d){in_delim}(\d{{4}})".format(
                in_delim=in_delim
            ),
            value,
        )
    elif output_format.startswith("Date (YYYY"):
        m = re.match(
            r"(\d{{4}}){in_delim}([0123]?\d)([01]?\d)".format(in_delim=in_delim), value
        )
    year = None
    month = None
    day = None
    if m:
        if output_format.startswith("Date (DD"):
            year = m.group(3)
            month = m.group(2).zfill(2)
            day = m.group(1).zfill(2)
        elif output_format.startswith("Date (MM"):
            year = m.group(3)
            month = m.group(1).zfill(2)
            day = m.group(2).zfill(2)
        elif output_format.startswith("Date (YYYY"):
            year = m.group(1)
            month = m.group(2).zfill(2)
            day = m.group(3).zfill(2)
    return (year, month, day)


def generate_output_date(year, month, day, output_format, out_delim):
    converted_value = ""
    if output_format.endswith(
        " to YYYY{out_delim}MM{out_delim}DD)".format(out_delim=out_delim)
    ):
        converted_value = "%s%s%s%s%s" % (year, out_delim, month, out_delim, day)
    if output_format.endswith(
        " to DD{out_delim}MM{out_delim}YYYY)".format(out_delim=out_delim)
    ):
        converted_value = "%s%s%s%s%s" % (day, out_delim, month, out_delim, year)
    if output_format.endswith(
        " to MM{out_delim}DD{out_delim}YYYY)".format(out_delim=out_delim)
    ):
        converted_value = "%s%s%s%s%s" % (month, out_delim, day, out_delim, year)
    return converted_value


def convert_date(value, output_format, idx_col, idx_row):
    converted_value = ""
    (in_delim, out_delim) = determine_date_delimiters(output_format)
    if in_delim == "" and len(value) != 8:
        logging.critical(
            "Invalid date value '%s' for format '%s' in field %d on row %d "
            "(ignoring the header), day and month must contain leading 0's. Exiting..."
            % (value, output_format, idx_col, idx_row)
        )
        sys.exit(24)

    (year, month, day) = parse_input_date(value, output_format, in_delim)

    # Is it a valid date?
    try:
        datetime.datetime.strptime("%s%s%s" % (year, month, day), "%Y%m%d")
    except ValueError:
        pass
    else:
        converted_value = generate_output_date(
            year, month, day, output_format, out_delim
        )
    if not converted_value:
        logging.critical(
            "Invalid date value '%s' for format '%s' in field %d on row %d "
            "(ignoring the header). Exiting..."
            % (value, output_format, idx_col, idx_row)
        )
        sys.exit(18)
    return converted_value


def convert_cell(value, output_format, idx_col, idx_row):
    if output_format not in SUPPORTED_OUTPUT_FORMATS:
        logging.critical(
            "Invalid output format '%s', must be one of '%s'. "
            "Exiting..."
            % (
                output_format,
                "', '".join(SUPPORTED_OUTPUT_FORMATS),
            )
        )
        sys.exit(27)
    converted_value = ""
    if "Keep numeric" == output_format:
        # Strip all non-numeric characters
        value = re.sub(r"\D", "", value)
    if (
        output_format in ("Integer", "Decimal", "Keep numeric")
        and str(value).strip() == ""
    ):
        value = "0"
    if "Time" == output_format:
        m = re.match(r"(\d{2})(:)?(\d{2})", value)
        if m:
            converted_value = "%s%s" % (m.group(1), m.group(3))
        else:
            logging.critical(
                "Invalid time format '%s' in field %d on row %d (ignoring the header). "
                "Exiting..." % (value, idx_col, idx_row)
            )
            sys.exit(17)
    elif output_format.startswith("Date ("):
        converted_value = convert_date(value, output_format, idx_col, idx_row)
    elif "Decimal" == output_format:
        # Decimal numbers must be sent with 2 decimal places and
        # *without* the decimal separator
        try:
            # Convert to string, then float respecting the user's Locale,
            # multiply by 100, round without decimals,
            # convert to integer (to drop the extra decimal values) then
            # finally back to string...
            converted_value = str(value)
            converted_value = atof(converted_value)
            converted_value = float(converted_value)
            converted_value = converted_value * 100
            converted_value = round(converted_value, 0)
            converted_value = int(converted_value)
            converted_value = str(converted_value)
        except ValueError:
            logging.critical(
                "Invalid decimal format '%s' in field %d on row %d (ignoring the "
                "header). Exiting..." % (value, idx_col, idx_row)
            )
            sys.exit(19)
    else:
        converted_value = value
    return converted_value


def convert_content(
    input_content, config, date_field_to_report_on=None, truncate=None, divert=None
):
    output_content = []
    diverted_output_content = []
    if date_field_to_report_on:
        # Argument is 1-based
        date_field_to_report_on -= 1
    oldest_date = "99999999"
    most_recent_date = "00000000"
    for idx_row, row in enumerate(input_content):
        converted_row_content = []
        divert_row = False
        # Confirm that the input_content doesn't have more fields than are
        # defined in the configuration file
        if len(row) > len(config):
            logging.critical(
                "Row %d (ignoring the header) has more fields than are defined in the "
                "configuration file! The row has %d fields while the configuration "
                "defines only %d possible fields. Exiting..."
                % (idx_row + 1, len(row), len(config))
            )
            sys.exit(23)
        for idx_col, cell in enumerate(row):
            output_format = config[idx_col]["output_format"]
            length = config[idx_col]["length"]

            if divert and idx_col + 1 in divert and str(cell) in divert[idx_col + 1]:
                # This field contains a value marked for content diversion
                # The content for the entire row will be diverted to a separate file
                divert_row = True

            if config[idx_col]["skip_field"]:
                cell = ""
            else:
                cell = convert_cell(cell, output_format, idx_col + 1, idx_row + 1)

            # Confirm that the length of the field (before padding) is less
            # than the maximum allowed length
            if len(cell) > config[idx_col]["length"]:
                if not truncate or idx_col + 1 not in truncate:
                    logging.critical(
                        "Field %d on row %d (ignoring the header) is too long! Length: "
                        "%d, max length %d. Exiting..."
                        % (
                            idx_col + 1,
                            idx_row + 1,
                            len(cell),
                            config[idx_col]["length"],
                        )
                    )
                    sys.exit(20)
                else:
                    # Truncate to the defined maximum field length
                    logging.info(
                        "Field %d on row %d (ignoring the header) is too long! Length: "
                        "%d, max length %d. Truncating field to its max length."
                        % (
                            idx_col + 1,
                            idx_row + 1,
                            len(cell),
                            config[idx_col]["length"],
                        )
                    )
                    cell = cell[: config[idx_col]["length"]]

            padded_output_value = pad_output_value(cell, output_format, length)
            converted_row_content.append(padded_output_value)

            if date_field_to_report_on and date_field_to_report_on == idx_col:
                if padded_output_value < oldest_date:
                    oldest_date = padded_output_value
                if padded_output_value > most_recent_date:
                    most_recent_date = padded_output_value
        # Process fields not in the input content but defined in the
        # configuration file: empty padding, based on the defined output format
        for idx_col in range(len(row), len(config)):
            output_format = config[idx_col]["output_format"]
            length = config[idx_col]["length"]
            padded_output_value = pad_output_value("", output_format, length)
            converted_row_content.append(padded_output_value)
        if divert_row:
            diverted_output_content.append("".join(converted_row_content))
        else:
            output_content.append("".join(converted_row_content))

    logging.debug("The output content:\n%s" % "\n".join(output_content))
    return (output_content, diverted_output_content, oldest_date, most_recent_date)


def read_input_file(
    input_file, delimiter, quotechar, skip_header, skip_footer, encoding
):
    content = None
    with codecs.open(input_file, "r", encoding) as csvfile:
        content = csv.reader(csvfile, delimiter=delimiter, quotechar=quotechar)
        # Skip the header and footer if necessary
        content = list(content)
        num_lines = len(content)
        content = content[skip_header : num_lines - skip_footer]
        logging.debug(
            "There are %d lines in the input file %s:" % (num_lines, input_file)
        )
        if skip_header > 0 or skip_footer > 0:
            logging.debug(
                "Skipping %d header and %d footer lines" % (skip_header, skip_footer)
            )
        for row in content:
            logging.debug(" ||| ".join(row))
    return content


def load_config(config_file):
    config = []
    supported_skip_field = ("True", "False", "", None)
    logging.debug("Loading configuration %s" % config_file)

    # Open the configuration file (an Excel .xlsx file)
    wk = load_workbook(filename=config_file)
    ws = wk.active  # Get active worksheet or wk['some_worksheet']

    # Analyze the header to identify the relevant columns
    length_col = -1
    output_format_col = -1
    skip_field_col = -1
    for row in ws.iter_rows(max_row=1):
        for idx, cell in enumerate(row):
            if "Length" == cell.value:
                length_col = idx
            elif "Output format" == cell.value:
                output_format_col = idx
            elif "Skip field" == cell.value:
                skip_field_col = idx
    column_indices = (length_col, output_format_col, skip_field_col)
    if -1 in column_indices:
        logging.critical(
            "Invalid config file, missing one of the columns 'Length', 'Output format' "
            "or 'Skip field'. Exiting..."
        )
        sys.exit(13)

    # Loop over all the config rows (skipping the header)
    for idx_row, row in enumerate(ws.iter_rows(min_row=2)):
        config.append({})
        for idx_col, cell in enumerate(row):
            if idx_col == length_col:
                config[idx_row]["length"] = -1
                if isinstance(cell.value, int):
                    config[idx_row]["length"] = cell.value
                if isinstance(cell.value, str) and cell.value.isnumeric():
                    config[idx_row]["length"] = int(cell.value)
                if config[idx_row]["length"] < 0:
                    logging.critical(
                        "Invalid value '%s' for the 'Length' column on row %d, must be "
                        "a positive number. Exiting..." % (cell.value, idx_row + 2)
                    )
                    sys.exit(14)
            if idx_col == output_format_col:
                if cell.value in SUPPORTED_OUTPUT_FORMATS:
                    config[idx_row]["output_format"] = cell.value
                else:
                    logging.critical(
                        "Invalid output format '%s' on row %d, must be one of '%s'. "
                        "Exiting..."
                        % (
                            cell.value,
                            idx_row + 2,
                            "', '".join(SUPPORTED_OUTPUT_FORMATS),
                        )
                    )
                    sys.exit(15)
            if idx_col == skip_field_col:
                if cell.value in supported_skip_field:
                    config[idx_row]["skip_field"] = "True" == cell.value
                else:
                    logging.critical(
                        "Invalid value '%s' for the 'Skip field' column on row %d, "
                        "must be one  of 'True', 'False' or empty. Exiting..."
                        % (cell.value, idx_row + 2)
                    )
                    sys.exit(16)

    logging.info("Config '%s' loaded successfully" % config_file)
    logging.debug(config)
    return config


def get_version(rel_path):
    with open(rel_path) as f:
        for line in f.read().splitlines():
            if line.startswith("__version__"):
                return line.split('"')[1]
        else:
            raise RuntimeError("Unable to find version string.")


def validate_divert(divert):
    divert_values = {}
    for d in divert:
        v = d.split(",", 1)
        try:
            v[0] = int(v[0])
        except ValueError:
            logging.critical(
                "<field number> must be a number, as passed to the `--divert` "
                'argument in the format "<field number>,<value to divert on>" '
                "(without quotes). Exiting..."
            )
            sys.exit(28)
        if len(v) == 2:
            if v[0] not in divert_values:
                divert_values[v[0]] = []
            divert_values[v[0]].append(v[1])
        else:
            logging.critical(
                'The `--divert` argument must be formatted as "<field number>,'
                '<value to divert on>" (without quotes). Exiting...'
            )
            sys.exit(29)
    return divert_values


def validate_input_output_args(args):
    if args.input:
        if not os.path.isfile(args.input):
            logging.critical("The specified input file does not exist. Exiting...")
            sys.exit(10)
    elif args.input_directory:
        if not args.output_directory:
            logging.critical(
                "The `--output_directory` argument must be specified in addition to "
                "the `--input_directory` argument. Exiting..."
            )
            sys.exit(32)
        if not os.path.isdir(args.input_directory):
            logging.critical(
                "The value passed as `--input-directory` argument is not a valid "
                "folder path. Exiting..."
            )
            sys.exit(33)
        if not os.path.isdir(args.output_directory):
            pathlib.Path(args.output_directory).mkdir(parents=True, exist_ok=True)


def validate_shared_args(args):
    validate_input_output_args(args)
    if args.move_input_files and not args.output_directory:
        logging.critical(
            "The `--move-input-files` argument can only be used in combination with "
            "the `--output-directory` argument. Exiting..."
        )
        sys.exit(35)
    if not os.path.isfile(args.config):
        logging.critical("The specified configuration file does not exist. Exiting...")
        sys.exit(12)
    if args.skip_header != 0:
        try:
            args.skip_header = int(args.skip_header)
        except ValueError:
            logging.critical("The `--skip-header` argument must be numeric. Exiting...")
            sys.exit(21)
    if args.skip_footer != 0:
        try:
            args.skip_footer = int(args.skip_footer)
        except ValueError:
            logging.critical("The `--skip-footer` argument must be numeric. Exiting...")
            sys.exit(22)
    if args.truncate:
        truncate = []
        for t in args.truncate.split(","):
            try:
                truncate.append(int(t))
            except ValueError:
                logging.critical(
                    "The `--truncate` argument must be a comma-delimited list of "
                    "numbers. Exiting..."
                )
                sys.exit(25)
        args.truncate = truncate
    if args.divert:
        args.divert = validate_divert(args.divert)


def add_shared_args(parser):
    input_group = parser.add_mutually_exclusive_group(required=True)
    input_group.add_argument(
        "-i", "--input", help="Specify the input file", action="store"
    )
    input_group.add_argument(
        "-id",
        "--input-directory",
        help="Specify the input directory from which to process input files",
        action="store",
    )
    parser.add_argument(
        "-ie",
        "--input-encoding",
        help="Specify the encoding of the input files (default: 'utf-8')",
        action="store",
        required=False,
        default="utf-8",
    )
    output_group = parser.add_mutually_exclusive_group(required=True)
    output_group.add_argument(
        "-o", "--output", help="Specify the output file", action="store"
    )
    output_group.add_argument(
        "-od",
        "--output-directory",
        help="The directory in which to create the output files",
        action="store",
    )
    parser.add_argument(
        "-m",
        "--move-input-files",
        help="Move the input files to the output directory after processing. Must be "
        "used in conjunction with the `--output-directory` argument.",
        action="store_true",
        required=False,
    )
    parser.add_argument(
        "-c",
        "--config",
        help="Specify the configuration file",
        action="store",
        required=True,
    )
    parser.add_argument(
        "-dl",
        "--delimiter",
        help="The field delimiter used in the input file (default ,)",
        action="store",
        required=False,
        default=",",
    )
    parser.add_argument(
        "-q",
        "--quotechar",
        help='The character used to wrap textual fields in the input file (default ")',
        action="store",
        required=False,
        default='"',
    )
    parser.add_argument(
        "-sh",
        "--skip-header",
        help="The number of header lines to skip (default 0)",
        action="store",
        required=False,
        default=0,
    )
    parser.add_argument(
        "-sf",
        "--skip-footer",
        help="The number of footer lines to skip (default 0)",
        action="store",
        required=False,
        default=0,
    )
    parser.add_argument(
        "-l",
        "--locale",
        help="Change the locale, useful to handle decimal separators",
        action="store",
        required=False,
        default="",
    )
    parser.add_argument(
        "-t",
        "--truncate",
        help="Comma-delimited list of field numbers for which the output will be "
        "truncated at the maximum line length, should the input value be longer than "
        "the maximum defined field length. If not set, a field that is too long will "
        "cause the script to stop with an error.",
        action="store",
        required=False,
        default=[],
    )
    parser.add_argument(
        "-dv",
        "--divert",
        help="Diverts to a separate file the content from rows containing a specific "
        'value at a specific place. The format of this parameter is "<field number>'
        ',<value to divert on>" (without quotes). This parameter can be repeated '
        "several times to support different values or different fields. The diverted "
        "content will be saved to a file whose name will be the output filename with "
        '"_diverted" added before the file extension.',
        action="append",
        required=False,
        default=[],
    )


def parse_args(arguments):
    parser = argparse.ArgumentParser(
        description="Convert files from delimited (e.g. CSV) to fixed width format"
    )
    parser.add_argument(
        "--version",
        action="version",
        version="%s %s" % ("%(prog)s", get_version("__init__.py")),
    )

    parser.add_argument(
        "-x",
        "--overwrite-file",
        help="Allow to overwrite the output file",
        action="store_true",
        required=False,
    )

    add_shared_args(parser)

    parser.add_argument(
        "-d",
        "--debug",
        help="Print lots of debugging statements",
        action="store_const",
        dest="loglevel",
        const=logging.DEBUG,
        default=logging.WARNING,
    )
    parser.add_argument(
        "-v",
        "--verbose",
        help="Be verbose",
        action="store_const",
        dest="loglevel",
        const=logging.INFO,
    )
    args = parser.parse_args(arguments)

    # Configure logging level
    if args.loglevel:
        logging.basicConfig(level=args.loglevel)
        args.logging_level = logging.getLevelName(args.loglevel)

    # Validate if the arguments are used correctly
    if args.output and os.path.isfile(args.output) and not args.overwrite_file:
        logging.critical(
            "The specified output file does already exist, will NOT overwrite. Add "
            "the `--overwrite-file` argument to allow overwriting. Exiting..."
        )
        sys.exit(11)
    if args.overwrite_file and not args.output:
        logging.critical(
            "The `--overwrite-file` argument can only be used in combination with "
            "the `--output` argument. Exiting..."
        )
        sys.exit(34)
    if args.input and not args.output:
        logging.critical(
            "The `--output` argument must be specified in addition to the "
            "`--input` argument. Exiting..."
        )
        sys.exit(31)
    validate_shared_args(args)

    logging.debug("These are the parsed arguments:\n'%s'" % args)
    return args


def process(
    input,
    output,
    config,
    delimiter,
    quotechar,
    skip_header,
    skip_footer,
    date_field_to_report_on=None,
    locale="",
    truncate=None,
    divert=None,
    input_encoding="utf-8",
):
    # By default, set to the user's default locale, used to appropriately handle
    # Decimal separators
    setlocale(LC_NUMERIC, locale)

    define_supported_output_formats()

    config = load_config(config)

    if truncate:
        for t in truncate:
            if t > len(config):
                logging.critical(
                    "The value %d passed in the `--truncate` argument is invalid, it "
                    "is higher than the %d fields defined in the configuration file. "
                    "Exiting..." % (t, len(config))
                )
                sys.exit(26)
    if divert:
        for d in divert.keys():
            if d > len(config):
                logging.critical(
                    "The value %d passed as field ID in the `--divert` argument is "
                    "invalid, it is higher than the %d fields defined in the "
                    "configuration file. Exiting..." % (d, len(config))
                )
                sys.exit(30)

    input_content = read_input_file(
        input, delimiter, quotechar, skip_header, skip_footer, input_encoding
    )

    (
        output_content,
        diverted_output_content,
        oldest_date,
        most_recent_date,
    ) = convert_content(
        input_content, config, date_field_to_report_on, truncate, divert
    )

    write_output_file(output_content, output)
    if diverted_output_content:
        # Save the diverted content to its separate file with "_diverted" added before
        # the extension
        diverted_output = "%s_diverted%s" % (os.path.splitext(output))
        write_output_file(diverted_output_content, diverted_output)

    return (len(input_content), oldest_date, most_recent_date)


def init():
    if __name__ == "__main__":
        # Parse the provided command-line arguments
        args = parse_args(sys.argv[1:])

        input_output_files = []
        if args.input:
            # Just a single input/output files combination
            input_output_files = [(args.input, args.output)]
        elif args.input_directory:
            # Process all files in that top-level directory (no subdirectories)
            (_, _, filenames) = next(os.walk(args.input_directory))
            for ifile in filenames:
                input_file = os.path.join(args.input_directory, ifile)
                output_file = os.path.join(
                    args.output_directory, "%s_processed%s" % (os.path.splitext(ifile))
                )
                input_output_files.append((input_file, output_file))

        for (input_file, output_file) in input_output_files:
            logging.info("Processing input file %s", input_file)
            process(
                input_file,
                output_file,
                args.config,
                args.delimiter,
                args.quotechar,
                args.skip_header,
                args.skip_footer,
                None,
                args.locale,
                args.truncate,
                args.divert,
                args.input_encoding,
            )
            if args.move_input_files:
                shutil.move(input_file, args.output_directory)


init()
