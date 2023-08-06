"""A simple API to store/load python objects to/from spreadsheets

Limitations
-----------
Currently only classes with 51 or less attributes are supported

Notes
-----

- xlsx writing is significantly faster than csv currently
- When readable flag is set:
    - values are deserailized as strings not their original types (list, tuple, dict etc.)
    - Since csv is auto-interpreted for newline termination characters in most applications tabs are used as delimiters instead on csv files

Examples
--------
### Store some animal instances in a spreadsheet called 'animals.xlsx'
```
from ezspreadsheet import Spreadsheet

class Animal():
    def __init__(self, name:str, conservation_status:str):
        self.name = name
        self.conservation_status = conservation_status

leopard_gecko = Animal('Leopard Gecko', 'Least Concern')

philippine_eagle = Animal('Philippine Eagle', 'Threatened')

with Spreadsheet('animals.xlsx', Animal) as output_sheet:
    output_sheet.store(leopard_gecko, philippine_eagle)
```

### Store a list of instances into a spreadsheet called 'users.csv'
```
from ezspreadsheet import Spreadsheet

import random
import string
from dataclasses import dataclass

@dataclass
class User():
    Name:str
    Age:int
    Weight:int
    Family: list # Note that Iterables will be flattened to a string with newline seperators

instances = []
ranstring = lambda: ''.join(random.choices(string.ascii_uppercase, k=10)) # Generates a random 10 character string
for i in range(1000):
    instances.append(User(ranstring(), random.randint(12,100), random.randint(75,400), [ranstring(), ranstring(), ranstring()]))

with Spreadsheet('users.csv', User) as output_sheet:
    output_sheet.store(instances)
```

### Read the values back from the spreadsheet in example 1
```
from ezspreadsheet import Spreadsheet

class Animal():
    def __init__(self, name:str, conservation_status:str):
        self.name = name
        self.conservation_status = conservation_status

leopard_gecko = Animal('Leopard Gecko', 'Least Concern')

philippine_eagle = Animal('Philippine Eagle', 'Threatened')

with Spreadsheet('animals.xlsx') as output_sheet: # Returned values will be a namedtuple called Animal
    returned_class, instances = output_sheet.load("Animal")

for instance in instances:
    print(instance) \"\"\"prints Animal(name='Leopard Gecko', conservation_status='Least Concern')\nAnimal(name='Philippine Eagle', conservation_status='Threatened')\"\"\"

with Spreadsheet('animals.xlsx', Animal) as output_sheet: # Returned values will be an Animal class
    returned_class, instances = output_sheet.load("Animal")

for instance in instances:
    print(vars(instance)) \"\"\"prints: {'name': 'Leopard Gecko', 'conservation_status': 'Least Concern'}\n{'name': 'Philippine Eagle', 'conservation_status': 'Threatened'}\"\"\"
```
"""
import csv                                   # Used to read and write to CSV files
import logging                               # Used to log data for debugging and transparency
import datetime                              # Used to validate type assertions for datetime instances
from collections import namedtuple           # Used to deserailize classes from spreadsheets and instances
from typing import Any, Union, Iterable      # Used for type hinting and type assertions on various class methods

# Third party dependencies
import colored                               # Colours terminal output for emphasis
from openpyxl import Workbook, load_workbook # Used to open and operate with xlsx files
from openpyxl.styles import Font, Alignment  # Used to style various output to xlsx files


class Spreadsheet():
    """A class that allows serialization/deserialization of python objects to csv or xlsx files

    Parameters
    ----------
    file_name : (str)
        The name of the .xlsx or .csv file that will be saved out to or loaded in

    class_identifier : (object or bool)
        The class object for instances you want to store, see example(s) for details
        If not specified (left as False), it's assumed you only want to load values

    Raises
    ------
    ValueError

        In three cases:

            1. If instances provided to Spreadsheet.store() do not match type used to construct Spreadsheet instance
            2. If class provided has more than 51 attributes (see limitations section of docs for details)
            3. The provided file is not a .xlsx file or a .csv file

    Examples
    --------
    ## Store some animal instances in a spreadsheet called 'animals.xlsx'
    ```
    from ezspreadsheet import Spreadsheet

    class Animal():
        def __init__(self, name:str, conservation_status:str):
            self.name = name
            self.conservation_status = conservation_status

    leopard_gecko = Animal('Leopard Gecko', 'Least Concern')

    philippine_eagle = Animal('Philippine Eagle', 'Threatened')

    with Spreadsheet('animals.xlsx', Animal) as output_sheet:
        output_sheet.store(leopard_gecko, philippine_eagle)
    ```
    """
    def __init__(self, file_name:str, class_identifier:object=False):
        if file_name.endswith(".xlsx"):
            self.spreadsheet = _XLSX_Spreadsheet(file_name, class_identifier)
        elif file_name.endswith(".csv"):
            self.spreadsheet = _CSV_Spreadsheet(file_name, class_identifier)
        else:
            raise ValueError(f"Provided file {file_name} is not a csv or xlsx file")

    def __enter__(self):
        """entrypoint for the context manager"""
        return self.spreadsheet.__enter__()

    def __exit__(self, exc_type, exc_value, traceback):
        """Exitpoint for the context manager

        Returns
        -------
        bool
            True if the context manager ran into no issues saving files
        """
        return self.spreadsheet.__exit__(exc_type, exc_value, traceback)

    def store(self, *instances:Union[object, Iterable[object]], readable:bool = False):
        """Takes in instance(s) of the specified class to store

        Parameters
        ----------
        instances : (Iterable[object] or arbitrary number of isntances)
            The instances with the data you want to store

        readable : bool
            If True iterable attributes are written as readable values instead of directly storing iterables, by default False

        Notes
        -----

        - iterables stored while readable == true cannot be deserialized to their original type

        Raises
        ------
        ValueError
            If an instance is not the correct type

        Notes
        -----

        - Any methods are not serialized, only attribtues

        Examples
        --------
        ## Store some animal instances in a spreadsheet called 'animals.xlsx'
        ```
        from ezspreadsheet import Spreadsheet

        class Animal():
            def __init__(self, name:str, conservation_status:str):
                self.name = name
                self.conservation_status = conservation_status

        leopard_gecko = Animal('Leopard Gecko', 'Least Concern')

        philippine_eagle = Animal('Philippine Eagle', 'Threatened')

        with Spreadsheet('animals.xlsx', Animal) as output_sheet:
            output_sheet.store(leopard_gecko, philippine_eagle)
        ```
        """
        return self.spreadsheet.store(*instances, readable=readable)

    def load(self, name:str) -> tuple:
        """Loads the class, and instances stored inside Spreadsheet at self.file_name

        Parameters
        ----------
        name : str
            The name you want to assign the class that is returned

        Notes
        -----

        - if self.class_identifier is specified on Spreadsheet instantiation then that class is used instead of instantiating a new one
        - if self.class_identifier is not specified a subclass of namedtuple is instantiated and passed back

        Returns
        -------
        tuple
            First return value is the constructor used to create instances (class if class_identifier is specified, else custom derived class), and second all the found instances

        Raises
        ------
        ValueError
            If file is empty, or header of file is not equivalent to provided class attributes
        
        Notes
        -----

        - If you didn't specify a class identifier when opening the spreadsheet the returned values are namedtuples and not full class instances

        Examples
        --------
        ## Loading some stored values of the Animal class from animals.xlsx
        ```
        with Spreadsheet('animals.xlsx') as loaded_sheet:
            Animal, instances = loaded_sheet.load('Animal')

        # NOTE: Animal at this point is a namedtuple constructor, not a full python class

        print(Animal) # Prints: <class '__main__.Animal'>
        print(instances) # Prints: [Animal(name='Leopard Gecko', conservation_status='Least Concern'), Animal(name='Philippine Eagle', conservation_status='Threatened')]
        ```

        ## Loading some stored values of the Animal class from animals.xlsx with the class identifier specified
        ```
        class Animal():
            def __init__(self, name:str, conservation_status:str):
                self.name = name
                self.conservation_status = conservation_status
        
        with Spreadsheet('animals.xlsx', Animal) as loaded_sheet:
            Animal, instances = loaded_sheet.load('Animal')
    
        print(Animal) # Prints: <class '__main__.Animal'>

        for instance in instances:
            print(vars(instance)) # Since these are real class instances we can use vars()
        '''prints:
        {'name': 'Leopard Gecko', 'conservation_status': 'Least Concern'}
        {'name': 'Philippine Eagle', 'conservation_status': 'Threatened'}
        '''
        ```
        """
        self.spreadsheet.load(name)


    def _get_values_from_instance(self, instance:object) -> list:
        """Get's the instance's attribute values

        Parameters
        ----------
        instance : object
            The instance to pull the attribute values from

        Returns
        -------
        list
            The values for the attributes from the instance
        """
        logging.debug(f"Attributes are {self.class_attributes}")
        values = [] # All the values of the attributes in order
        for attribute in self.class_attributes:
            logging.debug(f"Looking for attribute {attribute} found value {instance.__dict__[attribute]}")
            values.append(instance.__dict__[attribute]) 
        return values


class _XLSX_Spreadsheet(Spreadsheet):
    """A class that takes in instances of objects and serializes them to xlsx files

    Parameters
    ----------
    file_name : (str)
        The name of the .xlsx file that will be saved out (extension can be included or excluded)

    class_identifier : (object or bool)
        The class object for instances you want to store, see example(s) for details
        If not specified (left as False), it's assumed you only want to load values

    Raises
    ------
    ValueError

        In two cases:

            1. If instances provided to Spreadsheet.store() do not match type used to construct Spreadsheet instance
            2. If class provided has more than 51 attributes (see limitations section of docs for details)
    """
    def __init__(self, file_name:str, class_identifier:object=False):
        self.file_name = file_name
        self.workbook = None
        self.worksheet = None
        self.class_identifier = class_identifier
        self.class_attributes = None

        if class_identifier:
            # Get all attributes of class defined in __init__
            self.class_attributes = class_identifier.__init__.__code__.co_varnames[1::] # Skip the self
            if len(self.class_attributes) > 51:
                raise ValueError(f"Provided class {class_identifier.__name__} has more than 51 attributes")


    def __enter__(self):
        """Entrypoint for the context manager

        Returns
        -------
        Spreadsheet
            Reference to self
        """
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.page_setup.fitToWidth = 1
        return self


    def __exit__(self, exc_type, exc_value, traceback):
        """Exitpoint for the context manager

        Returns
        -------
        bool
            True if the context manager ran into no issues saving files
        """
        if exc_type is None and exc_value is None:
            try:
                self.workbook.save(self.file_name)
                print(f"{self.file_name} successfully saved")
            except PermissionError:
                input(f"{colored.fg(1)}File {self.file_name} is currently open{colored.fg(15)}\nPlease close it and hit enter to save file: ")
                self.workbook.save(self.file_name)
            return True

        else:
            print(f"{colored.fg(1)}Ran into exception {exc_type.__name__}, with value {exc_value} here is traceback{colored.fg(15)}")
            return False


    def _add_row_to_spreadsheet(self, data:list, row:int, style:Font = False, readable:bool = False):
        """Take in some data, and an int for a row and add that data to that row

        Parameters
        ----------
        data : list
            The data you want to store in the row

        row : int
            The index of the row to store the data to

        style : Font, optional
            If you want to supply custom formatting for the row, by default False

        readable : bool
            If True iterable attributes are written as readable values instead of directly storing iterables, by default False
        """
        # The value that will be converted using chr() for column identifiers i.e. A1 B1 etc.
        column_identifier = 65  # Initialize to ord() value of 'A'

        for value in data:
            if column_identifier == 91:  # Roll over to Ax column identifiers from x column identifiers
                label = f"AA{row}"
            elif column_identifier > 91:  # If beyond Z in column identifiers
                label = f"A{chr(column_identifier-26)}{row}"
            else:  # If before or at Z in column identifiers
                label = f"{chr(column_identifier)}{row}"
            logging.debug(f"{value} will be written to {label}")

            # Apply styles if specified
            if style:
                self.worksheet[label].font = style

            # Add value to worksheet
            if type(value) not in [str, int, float, datetime.datetime]:
                if type(value) == dict and readable:
                    print("Serializing dictionary in readable format") # TODO: remove
                    logging.debug("Serializing dictionary in readable format")
                    flattened_value = ""
                    for key in value:
                        flattened_value += f"- {key}: {value[key]}\n"
                    self.worksheet[label] = flattened_value
                
                elif readable:
                    # If value is an Iterable that's not a str, int or float then flatten it to a str
                    logging.debug(f"Serializing {type(value)} in readable format")
                    flattened_value = ""
                    for sub_value in value: 
                        flattened_value += f"- {str(sub_value)}\n"
                    self.worksheet[label] = flattened_value

                else:
                    # Value is not a str, int, float or datetime object (all can be natively serialized)
                    self.worksheet[label] = str(value)
            else: # If value is a string, int, float or datetime object
                self.worksheet[label] = value

            # Apply wrap text formatting to all rows that aren't the heading
            if not row == 1: 
                self.worksheet[label].alignment = Alignment(wrapText=True)

            # Increment the column identifiers variable to move to next column letter
            column_identifier += 1


    def store(self, *instances:Union[object, Iterable[object]], readable:bool = False):
        """Takes in instance(s) of the specified class to store

        Parameters
        ----------
        instances : (Iterable[object] or arbitrary number of isntances)
            The instances with the data you want to store

        readable : bool
            If True iterable attributes are written as readable values instead of directly storing iterables, by default False

        Raises
        ------
        ValueError
            If an instance is not the correct type
        """
        print(f"Beginning to store {self.class_identifier.__name__} instances to {self.file_name}")
        current_row = 1  # The current row that the iteration is at

        # Add heading with the list of class attributes to A1
        if not self.class_attributes:
            raise ValueError("No class constructor provided, cannot store instances")
        self._add_row_to_spreadsheet(self.class_attributes, current_row, Font(bold=True, size=14))
        current_row += 1  # Increment row to start with row right after heading
        logging.debug(f"Instances are {instances}")

        # Check if instance provided is a class of correct type, or an Iterable
        for current_instance in instances:
            logging.debug(f"Instance is {str(current_instance)}")
            if isinstance(current_instance, Iterable):  # If argument is an Iterable (i.e. list, tuple etc.)
                for sub_instance in current_instance:
                    if not isinstance(sub_instance, self.class_identifier):  # Validate sub-instance is correct type
                        raise ValueError(f"Provided instance: {sub_instance} is not of type {self.class_identifier}")
                    else:
                        self._add_row_to_spreadsheet(self._get_values_from_instance(sub_instance), current_row, readable=readable)
                        current_row += 1
            elif not isinstance(current_instance, self.class_identifier):  # If argument is not correct type
                raise ValueError(f"Provided instance: {current_instance} is not of type {self.class_identifier}")
            
            else:  # If argument is a single class instance of the correct type
                logging.debug(f"Adding values from {str(current_instance)}: {self._get_values_from_instance(current_instance)}")
                self._add_row_to_spreadsheet(self._get_values_from_instance(current_instance), current_row, readable=readable)
                current_row += 1

    def _load_values(self) -> list:
        """Yields each row of values to be consumed inside self.load()

        Yields
        -------
        list
            The values for a given row
        """
        for values in self.worksheet.values:
            values = list(values)
            for index, value in enumerate(values):
                # Deserialize iterables like lists, tuples and dicts
                if type(value) == str:
                    if value.startswith("["): # Deserialize lists that were not stored with self.store(readable=True)
                        value = value[1:-2].replace("\'", "").replace('\"', "").split(',')
                        values[index] = [v.strip() for v in value]

                    elif value.startswith("("): # Deserialize tuples that were not stored with self.store(readable=True)
                        value = value[1:-2].replace("\'", "").replace('\"', "").split(',')
                        values[index] = tuple(v.strip() for v in value)

                    elif value.startswith("{"): # Deserialize dicts that were not stored with self.store(readable=True)
                        key_value_pairs = value[1:-2].replace("\'", "").replace('\"', "").split(',')
                        result = {}
                        for pair in key_value_pairs:
                            key, value = pair.split(":")
                            key = key.strip()
                            if type(value) == str:
                                result[key] = value.strip()
                            else:
                                result[key] = value
                        values[index] = result
            yield values

    def load(self, name:str) -> tuple:
        """Loads the class, and instances stored inside Spreadsheet at self.file_name

        Parameters
        ----------
        name : str
            The name you want to assign the class that is returned

        Returns
        -------
        tuple
            First return value is the constructor used to create instances (class if class_identifier is specified, else namedtuple), and second all the found instances
        
        Raises
        ------
        ValueError
            In 2 cases:

                1. If spreadsheet file is empty
                2. If header does not match provided class attributes
        """
        self.workbook = load_workbook(self.file_name)
        self.worksheet = self.workbook.active

        values = self._load_values()

        instances = []

        try:
            header = tuple(next(values)) # skip the attributes
        except StopIteration:
            raise ValueError("Provided spreadsheet is empty")

        if self.class_identifier: # If class was specified
            logging.debug(f"Class identifier {self.class_identifier} specified")

            if self.class_attributes != header: # Validate the file header is the same as the attributes provided
                raise ValueError(f"Provided header {header} is not the same as class attribues {self.class_attributes}")

            constructor = self.class_identifier
            for instance_values in values:
                instances.append(self.class_identifier(*instance_values))
        else:
            logging.debug("No class identifier specified, generating namedtuple")
            # Get attributes from first row
            constructor = namedtuple(name, header)

            for instance_values in values:
                instances.append(constructor._make(instance_values))

        logging.debug(f"Returning: {constructor}\n\n{instances}")
        return constructor, instances



class _CSV_Spreadsheet(Spreadsheet):
    def __init__(self, file_name:str, class_identifier:object=False):
        self.file_name = file_name
        self.spreadsheet_file = None
        self.reader = None
        self.writer = None
        self.read = False
        self.written = False
        self.class_identifier = class_identifier
        self.class_attributes = None

        if class_identifier:
            # Get all attributes of class defined in __init__
            self.class_attributes = class_identifier.__init__.__code__.co_varnames[1::] # Skip the self
            if len(self.class_attributes) > 51:
                raise ValueError(f"Provided class {class_identifier.__name__} has more than 51 attributes")

    def __enter__(self):
        """entrypoint for the context manager"""
        try:
            self.spreadsheet_file = open(self.file_name, 'r+', newline='\n')
        except PermissionError:
            input(f"{colored.fg(1)}File {self.file_name} is currently open{colored.fg(15)}\nPlease close it and hit enter to open file: ")
            self.spreadsheet_file = open(self.file_name, 'r+', newline='\n')
        except FileNotFoundError:
            temp = open(self.file_name, 'w+', newline='\n')
            temp.close()
            self.spreadsheet_file = open(self.file_name, 'r+', newline='\n')

        return self

    def __exit__(self, exc_type, exc_value, traceback):
        if exc_type is None and exc_value is None:
            try:
                self.spreadsheet_file.close()
                if self.written:
                    print(f"{self.file_name} successfully saved")
                if self.read:
                    print(f"{self.file_name} successfully read")
            except PermissionError:
                input(f"{colored.fg(1)}File {self.file_name} is currently open{colored.fg(15)}\nPlease close it and hit enter to save file: ")
                self.spreadsheet_file.close()
            return True

        else:
            print(f"{colored.fg(1)}Ran into exception {exc_type.__name__}, with value {exc_value} here is traceback{colored.fg(15)}")
            return False

    def store(self, *instances:Union[object, Iterable[object]], readable:bool = False):
        """Takes in instance(s) of the specified class to store

        Parameters
        ----------
        instances : (Iterable[object] or arbitrary number of isntances)
            The instances with the data you want to store

        readable : bool
            If True iterable attributes are written as readable values instead of directly storing iterables, by default False

        Raises
        ------
        ValueError
            If an instance is not the correct type, or no class constructor is provided
        """

        if not self.writer:
            self.writer = csv.writer(self.spreadsheet_file, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)

        # Write header
        if not self.class_attributes:
            raise ValueError("No class constructor provided, cannot store instances")
        self.writer.writerow(self.class_attributes)

        all_instance_values = [] # A list that will contain just the values of all instances

        for instance in instances: # Grab each instances' attribute values and store them in all_instance_values
            logging.debug(f"Instance is {str(instance)}")
            if isinstance(instance, Iterable):  # If argument is an Iterable (i.e. list, tuple etc.)
                for sub_instance in instance:
                    if not isinstance(sub_instance, self.class_identifier):  # Validate sub-instance is correct type
                        raise ValueError(f"Provided instance: {sub_instance} is not of type {self.class_identifier}")
                    else:
                        instance_values = self._get_values_from_instance(sub_instance)
                        all_instance_values.append(instance_values)
            elif not isinstance(instance, self.class_identifier):  # If instance is not correct type
                raise ValueError(f"Provided instance: {instance} is not of type {self.class_identifier}")

            else: # If instance provided is a single instance of correct type
                instance_values = self._get_values_from_instance(instance)
                all_instance_values.append(instance_values)

        if readable: # Write iterables as readable forms if flag is specified
            for instance_values in all_instance_values:
                for index, value in enumerate(instance_values):
                    if isinstance(value, Iterable): # If the current instance attribute is an iterable
                        if type(value) not in [str, int, float, datetime.datetime]:
                            if type(value) == dict and readable:
                                logging.debug("Serializing dictionary in readable format")
                                flattened_value = ""
                                for key in value:
                                    flattened_value += f"- {key}: {value[key]} \t"
                                instance_values[index] = flattened_value

                            else:
                                # If value is an Iterable that's not a str, int or float then flatten it to a str
                                logging.debug(f"Serializing {type(value)} in readable format")
                                flattened_value = ""
                                for sub_value in value:
                                    flattened_value += f"- {str(sub_value)} \t"
                                    instance_values[index] = flattened_value
        self.written = True
        self.writer.writerows(all_instance_values)


    def _load_values(self):
        for row in self.reader:
            values = list(row)
            for index, value in enumerate(values):
                # Deserialize iterables like lists, tuples and dicts
                if type(value) == str:
                    if value.startswith("["): # Deserialize lists that were not stored with self.store(readable=True)
                        value = value[1:-2].replace("\'", "").replace('\"', "").split(',')
                        values[index] = [v.strip() for v in value]

                    elif value.startswith("("): # Deserialize tuples that were not stored with self.store(readable=True)
                        value = value[1:-2].replace("\'", "").replace('\"', "").split(',')
                        values[index] = tuple(v.strip() for v in value)

                    elif value.startswith("{"): # Deserialize dicts that were not stored with self.store(readable=True)
                        key_value_pairs = value[1:-2].replace("\'", "").replace('\"', "").split(',')
                        result = {}
                        for pair in key_value_pairs:
                            key, value = pair.split(":")
                            key = key.strip()
                            if type(value) == str:
                                result[key] = value.strip()
                            else:
                                result[key] = value
                        values[index] = result

                    elif value.isdigit():
                        values[index] = int(value)

                    elif value.isdecimal():
                        values[index] = float(value)

                    elif value.startswith("-"): # possible negative integer
                        if value[1::].isdigit():
                            values[index] = int(value[1::]) * -1
                        if value[1::].isdecimal():
                            values[index] = float(value[1::]) * -1

            yield values


    def load(self, name:str) -> tuple:
        """Loads values from provided spreadsheet file

        Parameters
        ----------
        name : str
            The name you want to give to the returned class if no class constructor is provided

        Returns
        -------
        tuple
            First the class constructor, second the instances retrieved from the file

        Raises
        ------
        ValueError
            In 2 cases:

                1. If spreadsheet file is empty
                2. If header does not match provided class attributes
        """
        if not self.reader:
            self.reader = csv.reader(self.spreadsheet_file, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)

        values = self._load_values()
        
        try:
            header = next(values)
            logging.debug(f"Found header: {header}")
        except StopIteration:
            raise ValueError(f"File {self.file_name} is empty")
        if self.class_attributes:
            if not tuple(header) == self.class_attributes:
                raise ValueError(f"Header: {header} is not equivalent to privided class attributes: {self.class_attributes}")
        else:
            self.class_attributes = header

        instances = []

        if self.class_identifier: # If class was specified
            logging.debug(f"Class identifier {self.class_identifier} specified")
            constructor = self.class_identifier
            for instance_values in values:
                if len(instance_values) == len(self.class_attributes): # Sometimes a nonetype is deserialized
                    instances.append(self.class_identifier(*instance_values))
                elif instance_values:
                    print(f"Row {instance_values} was skipped")
        else:
            logging.debug("No class identifier specified, generating namedtuple")
            # Get attributes from first row
            base_named_tuple = namedtuple(name, header)
            class constructor(base_named_tuple):
                __dict__ = property(base_named_tuple._asdict)
                __name__ = property(base_named_tuple.__name__)

            constructor.__name__ = name

            for instance_values in values:
                if len(instance_values) == len(self.class_attributes):
                    instances.append(constructor._make(instance_values))

        logging.debug(f"Returning: {constructor}\n\n{instances}")
        self.read = True
        return constructor, instances


if __name__ == "__main__": # local test code to play around with
    class Animal():
        def __init__(self, name:str, conservation_status:str):
            self.name = name
            self.conservation_status = conservation_status
    
    leopard_gecko = Animal('Leopard Gecko', 'Least Concern')

    philippine_eagle = Animal('Philippine Eagle', 'Threatened')

    with Spreadsheet('animals.csv', Animal) as output_sheet:
        output_sheet.store(leopard_gecko, philippine_eagle)

    with Spreadsheet('animals.csv') as loaded_sheet:
        animals, instances = loaded_sheet.load('animals')
    