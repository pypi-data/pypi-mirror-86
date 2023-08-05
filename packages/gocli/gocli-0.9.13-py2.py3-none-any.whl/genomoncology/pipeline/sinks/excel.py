import os.path
import time

from cytoolz.curried import curry
from glom import glom
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.workbook import Workbook

from genomoncology.parse import DocType
from genomoncology.pipeline.converters import to_str
from .base import Sink

bold_font = Font(b=True)


@curry
class ExcelSink(Sink):
    def __init__(self, filename):
        self.filename = "output.xlsx" if filename == "-" else filename

        if os.path.exists(self.filename):
            self.wb = load_workbook(self.filename)
        else:
            self.wb = Workbook()

        try:
            del self.wb["Sheet"]
        except KeyError:
            pass

        self.sheets = {}
        self.columns = {}

    def write(self, unit):
        columns = DocType.get_default_columns(unit)
        sheet_name = DocType.get_sheet_name(unit)
        ws, is_new_sheet = self.get_sheet(sheet_name)

        if is_new_sheet:
            data = DocType.get_header_names(columns)
            ws.append(data)

        if not DocType.HEADER.is_a(unit):
            data = [self.get_item(unit, c) for c in columns]
            ws.append(data)

    def get_item(self, unit, column):
        value = glom(unit, column, default="")
        if isinstance(value, (list, tuple, set)):
            return ", ".join(map(to_str, value))
        else:
            return value

    def get_sheet(self, sheet_name):
        is_new_sheet = False

        if sheet_name not in self.sheets:
            self.sheets[sheet_name] = self.create_sheet(sheet_name)
            is_new_sheet = True

        return self.sheets[sheet_name], is_new_sheet

    def create_sheet(self, sheet_name):
        try:
            new_ws_title = f"{sheet_name}-{int(time.time())}"[:32]
            self.wb[sheet_name].title = new_ws_title
        except KeyError:
            pass

        return self.wb.create_sheet(sheet_name, 0)

    def close(self):
        for ws in self.sheets.values():
            for cell in ws["1:1"]:
                cell.font = bold_font

        self.wb.save(self.filename)
        self.wb.close()


#
