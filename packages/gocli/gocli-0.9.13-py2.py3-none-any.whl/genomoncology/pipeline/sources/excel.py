from genomoncology.parse.doctypes import DocType, __TYPE__
from cytoolz.curried import curry
from openpyxl import load_workbook

from .base import Source


def clean_row(row):
    for item in row:
        if isinstance(item, str):
            yield item.strip()
        else:
            yield item


@curry
class ExcelSource(Source):
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)

    def __iter__(self):
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            ws_data = ws.values
            columns = next(ws_data)

            for row in ws_data:
                record = dict(zip(columns, clean_row(row)))
                record["sheet_name"] = sheet_name
                record[__TYPE__] = DocType.TSV.value
                yield record
