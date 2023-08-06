from pathlib import Path
import openpyxl as xl
import pandas as pd
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import classification_report, accuracy_score
from sklearn.model_selection import train_test_split
import io
import requests
import os


class Model:
    """Making our Model"""

    def __init__(self):
        self.dataset = None
        self.model = None

    def load_train_data(self, dataset='orig_data'):

        if dataset == 'orig_data':
            print('############ Using Original Data for Model Training ############')
            #dataset_path = r'C:\Users\850047756\Desktop\Shortlister\Dataset\Dataset Builder_final - 17k.csv'
            url = "https://raw.githubusercontent.com/msp04/PyRecruiter/main/Dataset/Dataset%20Builder_final%20-%2017k.csv"
            s = requests.get(url).content
            self.dataset = pd.read_csv(io.StringIO(s.decode('utf-8')), sep=',')

        elif dataset == 'experience_data':
            print('############ Using Experience Data for Model Training ############')
            # dataset_path = r'https://raw.githubusercontent.com/mohitpawar473/PyRecruit/main/Dataset/Dataset%20Builder%20-%20Experience%2017k.csv'

            url = "https://raw.githubusercontent.com/msp04/PyRecruiter/main/Dataset/Dataset%20Builder%20-%20Experience%2017k.csv"
            p = requests.get(url).content
            self.dataset = pd.read_csv(io.StringIO(p.decode('utf-8')), sep=',')

        else:
            print('Dataset does not exist!')
            return

    def train_model(self, show_score=True, show_report=False):
        self.dataset['score'] = 2 * self.dataset['Story_telling_project'] + 2 * self.dataset['Domain_Project_Types']
        X = self.dataset.drop(['Shortlisting'], axis=1)
        y = self.dataset['Shortlisting']
        X_train, X_test, y_train, y_test = train_test_split(X.values, y.values, test_size=0.33, random_state=42)
        self.model = RandomForestClassifier(max_depth=10, class_weight='balanced', n_estimators=100, random_state=2)
        self.model.fit(X_train, y_train)

        pred = self.model.predict(X_test)
        pred_prob = self.model.predict_proba(X_test)[0]
        # print(pred,pred_prob*100)
        if show_score is True:
            print("\nYour Model Performance is {}%".format(accuracy_score(pred, y_test) * 100))
        if show_report is True:
            print(classification_report(pred, y_test))

    def load_test_data(self, data_path=None):
        if data_path is None:
            print("Please provide a data path")

        path = Path(data_path)

        if path.glob('*xls'):
            filename = path
            wb1 = xl.load_workbook(filename)
            ws1 = wb1.worksheets[0]

            # opening the destination excel file
            filename1 = r'C:\Users\mohit\PyRecruiter\pyrecruiter\file_template.xlsx'
            wb2 = xl.load_workbook(filename1)
            ws2 = wb2.active

            mr = ws1.max_row  # calculate total number of rows and
            mc = ws1.max_column  # columns in source excel file

            # copying the cell values from source
            # excel file to destination excel file
            for i in range(2, mr + 1):
                for j in range(1, mc + 1):
                    # reading cell value from source excel file
                    c = ws1.cell(row=i, column=j)

                    # writing the read value to destination excel file
                    ws2.cell(row=i, column=j).value = c.value

                    # saving the destination excel file
                    wb2.save(str('your_file.xlsx'))
            ws2.delete_rows(2, mr)
            wb2.save(filename1)
            print("\n>>>>>>>>>>>> Taking your Data for Prediction  >>>>>>>>>>>>")

    def evaluate(self):

        data = pd.read_excel('your_file.xlsx')
        data['score'] = 2 * data['Story_telling_project'] + 2 * data['Domain_Project_Types']
        final = data.drop(['name', 'email', 'Name', 'Email', 'Start time', 'Completion time', 'ID'], axis=1)
        prediction = self.model.predict(final)
        for i in range(0, len(final)):
            print("\nThe Predicted Result is {} and the Prediction Probability is as follows: Not Shortlisted "
                  "Probability is {}% and Shortlisted Probability is {}%.".format(prediction[i],
                                                                                  self.model.predict_proba(final)[i][
                                                                                      0] * 100,
                                                                                  self.model.predict_proba(final)[i][
                                                                                      1] * 100))

        print("\n############ Here 1 Stands for Shortlisted and 0 Stands for Not Shortlisted ################")
        print("\n>>>>>>>>>>>> Saving your Results in a csv File <<<<<<<<<<<<<<<<")
        new_df = pd.DataFrame(
            {'Name': data.name, 'Email': data.email, 'Prediction': prediction,
             'Not_Shortlisted_Probability': self.model.predict_proba(final)[i][0] * 100,
             'Shortlisted_Probability': self.model.predict_proba(final)[i][1] * 100})

        new_df.to_csv('prediction.csv', index=False)

